import time
import pandas as pd
import requests
import json
from openpyxl import load_workbook
from bs4 import BeautifulSoup
from difflib import get_close_matches
from datetime import datetime
import os

os.environ["OLLAMA_USE_CUDA"] = "1"  

Starting_time = datetime.now()
print(f"\n✅ Starting time: '{Starting_time}'")

MAX_RETRIES = 3
API_URL = "http://localhost:11434/api/generate"

# Load the Excel file
file_path = '/home/msaban/ChusCodification/small_codebook.xlsx'

# Load sheets using pandas
codes_sheet = pd.read_excel(file_path, sheet_name="Codes", header=None)
codif_sheet = pd.read_excel(file_path, sheet_name="Zero", header=None)

# Extract code definitions
definitions_mapping = {
    str(row[0]).strip().lower(): str(row[1]).strip() if not pd.isna(row[1]) else "No definition available"
    for row in codes_sheet.iloc[1:].values  # Skip header row
}

# Extract examples
examples_mapping = {
    str(row[0]).strip().lower(): str(row[2]).strip() if not pd.isna(row[2]) else "No example available"
    for row in codes_sheet.iloc[1:].values
}

# Get codes from "Codification" sheet (Columns F, G → indices 5,6)
code_columns = list(range(6, 7))
codes = [str(codif_sheet.iloc[0, col]).strip().lower() for col in code_columns]

# Function to find best match
def find_best_match(code_name, available_codes):
    matches = get_close_matches(code_name, available_codes, n=1, cutoff=0.7)
    return matches[0] if matches else None

# Fix codes mapping
fixed_codes = {
    raw_code: find_best_match(raw_code, definitions_mapping.keys()) or raw_code
    for raw_code in codes
}

# 🔹 Clean HTML content
def clean_html(html_text):
    return BeautifulSoup(str(html_text), "html.parser").get_text()

# Load workbook for writing results
workbook = load_workbook(file_path)
if "Zero" not in workbook.sheetnames:
    raise ValueError("❌ Sheet 'Zero' not found in the Excel file!")

workbook_sheet = workbook["Zero"]

# Define column indices
title_col = 0
category_col = 1
name_col = 2
description_col = 3  # Column D
embded_col = 4      # Column E

# Process each code column
for code_idx, code_col in enumerate(code_columns):
    raw_code_name = str(codif_sheet.iloc[0, code_col]).strip().lower()
    matched_code_name = fixed_codes.get(raw_code_name, raw_code_name)
    code_definition = definitions_mapping.get(matched_code_name, "No definition available")

    print(f"\n🚀 Processing Code: '{matched_code_name}'")
    print(f"📝 Definition: {code_definition}\n")

    # Reset Ollama context at the start of each column
    reset_data = {
        "model": "llama3.3:70b",
        "messages": [
            {"role": "system", "content": "Forget all previous instructions and start fresh."}
        ],
        "temperature": 0.0,
        "stream": False
    }
    requests.post(API_URL, headers={'Content-Type': 'application/json'}, json=reset_data)
    print(f"🧹 Ollama context cleared before processing column {code_col} ({matched_code_name})")

    # Process each row
    for i in range(1, 38):
        ils_title = codif_sheet.iloc[i, title_col]
        item_name = codif_sheet.iloc[i, name_col]
        item_category = codif_sheet.iloc[i, category_col]
        item_description = codif_sheet.iloc[i, description_col]
        item_embded_description = codif_sheet.iloc[i, embded_col]

        messages = [
            {"role": "system", "content": (
                f"You are a qualitative coding expert. You are assessing the student engagement of learning activities "
                f"created by teachers in an inquiry-based learning digital platform. These activities may have different "
                f"media content including text and embedded artifacts (e.g., images, videos, apps, labs). "
                f"Please review the provided activity description and code it based on the construct: `{matched_code_name}`. "
                f"The definition of this construct is `{code_definition}`. "
                f"After reviewing the text, assign a code of '1' if you believe the text exemplifies `{matched_code_name}`, "
                f"or a '0' if it does not. Your response should only be '1' or '0'."
            )},
            {"role": "user", "content": (
                f"Learning activity:\n"
                f"Lesson title: {clean_html(ils_title)}. \n"
                f"Activity category: {clean_html(item_category)}.\n"
                f"Activity name: {clean_html(item_name)}. \n"
                f"Activity description: {clean_html(item_description)}. \n"
                f"Embedded media content description: {clean_html(item_embded_description)} \n"
            )}
        ]

        data = {
            "model": "llama3.3:70b",
            "messages": messages,
            "temperature": 0.0,
            "stream": False
        }

        attempt = 0
        while attempt < MAX_RETRIES:
            try:
                response = requests.post(API_URL, headers={'Content-Type': 'application/json'}, json=data)
                response.raise_for_status()
                response_json = response.json()
                api_response = response_json.get("response", "").strip()
                result_value = api_response[0] if api_response and api_response[0] in ("1", "0") else "Error"
                print(f"📝 Row {i+1} - Code '{matched_code_name}': API response: {result_value}")
                workbook_sheet.cell(row=i+1, column=code_col+1, value=result_value)
                workbook.save(file_path)
                break
            except (requests.exceptions.RequestException, json.JSONDecodeError) as e:
                print(f"❌ Error for row {i+1}, code '{matched_code_name}': {e}")
                if attempt == MAX_RETRIES - 1:
                    workbook_sheet.cell(row=i+1, column=code_col+1, value="Error")
                    workbook.save(file_path)
            attempt += 1
            time.sleep(5)

workbook.save(file_path)
workbook.close()
Finishing_time = datetime.now()
print("\n✅ Results successfully written to the Excel file.")
print(f"\n✅ Starting time: '{Starting_time}'")
print(f"\n✅ Finishing time: '{Finishing_time}'")
