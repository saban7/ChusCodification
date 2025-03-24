import os
import time
import json
import requests
import pandas as pd
from bs4 import BeautifulSoup
from datetime import datetime
from openpyxl import load_workbook
from difflib import get_close_matches

# Enable GPU usage for Ollama (if supported)
os.environ["OLLAMA_USE_CUDA"] = "1"

# Constants
MAX_RETRIES = 3
API_URL = "http://localhost:11434/api/generate"
EXCEL_FILE_PATH = "/home/msaban/ChusCodification/Full/Context.xlsx"

def main():
    start_time = datetime.now()
    print(f"\n✅ Script started at: {start_time}")

    # Load the Excel file
    codes_sheet = pd.read_excel(EXCEL_FILE_PATH, sheet_name="Codes", header=None)
    codif_sheet = pd.read_excel(EXCEL_FILE_PATH, sheet_name="Codification", header=None)

    # Build dictionaries for definitions and examples from the "Codes" sheet
    definitions_mapping = {
        str(row[0]).strip().lower(): (str(row[1]).strip() if not pd.isna(row[1]) else "No definition available")
        for row in codes_sheet.iloc[1:].values  # Skip header
    }
    examples_mapping = {
        str(row[0]).strip().lower(): (str(row[2]).strip() if not pd.isna(row[2]) else "No example available")
        for row in codes_sheet.iloc[1:].values
    }

    # Identify the columns that contain codes in the "Zero" sheet (F, G → indices 5,6 in Excel → Python 0-based)
    code_columns = list(range(6, 19))  # Adjust if more columns should be processed
    # Extract the raw code names from the first row of these columns
    codes = [str(codif_sheet.iloc[0, col]).strip().lower() for col in code_columns]

    # Create a helper mapping from "raw_code" to "best_matched_code"
    available_codes = list(definitions_mapping.keys())
    fixed_codes = {raw: find_best_match(raw, available_codes) or raw for raw in codes}

    # Prepare to write results into the "Zero" sheet
    workbook = load_workbook(EXCEL_FILE_PATH)
    if "Codification" not in workbook.sheetnames:
        raise ValueError("❌ Sheet 'Codification' not found in the Excel file!")
    workbook_sheet = workbook["Codification"]

    # Column indices (0-based in pandas)
    title_col = 0       # Lesson title
    category_col = 1    # Activity category
    name_col = 2        # Activity name
    description_col = 3 # Activity description
    embed_col = 4       # Embedded media description
    summary_col = 5
    # Process each code column
    for code_col in code_columns:
        raw_code_name = str(codif_sheet.iloc[0, code_col]).strip().lower()
        matched_code_name = fixed_codes.get(raw_code_name, raw_code_name)
        code_definition = definitions_mapping.get(matched_code_name, "No definition available")
        code_example = examples_mapping.get(matched_code_name, "No example available")

        print(f"\n🚀 Processing Code: '{matched_code_name}'")
        print(f"📝 Definition: {code_definition}")
        print(f"📚 Example: {code_example}\n")

        # Loop through rows to generate and record codes
        for i in range(1, 758):  # Adjust range as needed
            # Read relevant information from the row
            lesson_title = codif_sheet.iloc[i, title_col]
            activity_category = codif_sheet.iloc[i, category_col]
            activity_name = codif_sheet.iloc[i, name_col]
            activity_description = codif_sheet.iloc[i, description_col]
            embed_description = codif_sheet.iloc[i, embed_col]
            previous_summary = codif_sheet.iloc[i, summary_col]
            # Construct the user text
            text_for_prompt = (
                f"Learning activity:\n"
                f"Activity description: {clean_html(activity_description)}.\n"
                f"Embedded media content description: {clean_html(embed_description)}.\n"
            )


            prompt = (
                f"You are a qualitative coding expert. You are assessing the student engagement of learning activities created by teachers in a inquiry-based learning digital platform. \n"
                f"These activities may have different media content including text and embedded artifacts (e.g., images, videos, apps, labs). Please review the provided activity description and code it based on the construct: `{matched_code_name}`. \n"
                f"The definition of this construct is `{code_definition}`.  \n"
                f"Here you have some examples: `{code_example}`. \n"
                f"For additional context, here is a summary of the 3 previous items: `{previous_summary}`. \n"
                f"After reviewing the text, assign a code of '1' if you believe the text exemplifies `{matched_code_name}`, or a '0' if it does not. Your response should only be '1' or '0'.\n\n"
                f"Text: `{text_for_prompt}`"
            )

            print(f"\n🤖 Ollama System Message: {prompt}\n")

            # Prepare data payload for the request
            data_payload = {
                "model": "llama3.3:70b",
                "prompt": prompt,
                "temperature": 0.0,
                "stream": False
            }

            # Attempt to get a valid response from Ollama
            result_value = send_to_ollama(data_payload, i, matched_code_name, workbook, workbook_sheet, code_col)
            
            print(f"📝 Row {i+1} - Code '{matched_code_name}': API response: {result_value}")
            # Write result to Excel
            workbook_sheet.cell(row=i+1, column=code_col+1, value=result_value)
            workbook.save(EXCEL_FILE_PATH)

            # Optional: A brief pause between requests
            # time.sleep(1)

    # Finalize
    workbook.save(EXCEL_FILE_PATH)
    workbook.close()

    end_time = datetime.now()
    print("\n✅ Results successfully written to the Excel file.")
    print(f"✅ Script started at: {start_time}")
    print(f"✅ Script finished at: {end_time}")

def find_best_match(code_name, available_codes):
    """
    Return the single closest match for 'code_name' from 'available_codes', 
    or None if no match exceeds the cutoff.
    """
    matches = get_close_matches(code_name, available_codes, n=1, cutoff=0.7)
    return matches[0] if matches else None

def clean_html(html_text):
    """
    Remove any HTML tags from the input text.
    """
    return BeautifulSoup(str(html_text), "html.parser").get_text()

def send_to_ollama(data_payload, row_idx, code_name, workbook, sheet, code_col):
    """
    Send 'data_payload' to the Ollama API, retrying up to MAX_RETRIES times.
    Returns a single-character response ('0' or '1') or 'Error' if unsuccessful.
    """
    for attempt in range(MAX_RETRIES):
        try:
            response = requests.post(API_URL, headers={'Content-Type': 'application/json'}, json=data_payload)
            response.raise_for_status()
            response_json = response.json()
            api_response = response_json.get("response", "").strip()

            # Validate Ollama response
            if api_response and api_response[0] in ("1", "0"):
                return api_response[0]  # Only store the first character
            else:
                # If Ollama returns something unexpected, log it
                print(f"⚠️ Unexpected response format for row {row_idx+1}, code '{code_name}': {api_response}")
                return "Error"

        except requests.exceptions.RequestException as req_err:
            print(f"❌ API connection error for row {row_idx+1}, code '{code_name}': {req_err}")
        except json.JSONDecodeError as json_err:
            print(f"⚠️ JSON decode error for row {row_idx+1}, code '{code_name}': {json_err}")

        # After a failed attempt, wait and retry
        if attempt < MAX_RETRIES - 1:
            time.sleep(5)
        else:
            # On final failure, write "Error" to Excel
            sheet.cell(row=row_idx+1, column=code_col+1, value="Error")
            workbook.save(EXCEL_FILE_PATH)
            print(f"🚨 Max retries reached for row {row_idx+1}. Moving to next item.")
            return "Error"

if __name__ == "__main__":
    main()
