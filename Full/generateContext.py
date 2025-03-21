import time
import pandas as pd
import requests
import json
from openpyxl import load_workbook
from bs4 import BeautifulSoup
from datetime import datetime

Starting_time = datetime.now()
print(f"\n✅ Starting time: '{Starting_time}'")

API_URL = "http://localhost:11434/api/generate"
EXCEL_FILE_PATH = "/home/msaban/ChusCodification/Full/Summary.xlsx"

# Load the Context sheet
codif_sheet = pd.read_excel(EXCEL_FILE_PATH, sheet_name="Codification", header=None)

# Load workbook for writing results
workbook = load_workbook(EXCEL_FILE_PATH)
if "Codification" not in workbook.sheetnames:
    raise ValueError("❌ Sheet 'Codification' not found in the Excel file!")

workbook_sheet = workbook["Codification"]

description_col = 3  # Column D
content_col = 4      # Column E
summary_col = 5      # Column F (where summaries will be stored)

# 🔹 Clean HTML content
def clean_html(html_text):
    return BeautifulSoup(str(html_text), "html.parser").get_text()

# 🔹 Generate summary using Llama
def generate_summary(last_descriptions, last_contents):
    if not last_descriptions and not last_contents:
        return "No previous context available."
    
    summary_text = " | ".join([f"Item{i}.Task description: {desc}\nItem{i}.embedded_artifact_description: {cont}" 
                               for i, (desc, cont) in enumerate(zip(last_descriptions, last_contents), start=1)])
    
    summary_prompt = (
        "Provide a summary of the instructions provided to the students and the embedded artifacts in the following items. "
        "The text should not take more than 50 words.\n\n"
        f"Text: `{summary_text}`"
    )
    
    data = {
        "model": "llama3.3:70b",
        "prompt": summary_prompt,
        "temperature": 0.0,
        "stream": False
    }
    
    response = requests.post(API_URL, headers={'Content-Type': 'application/json'}, json=data)
    
    if response.status_code == 200:
        return response.json().get("response", "No summary available.").strip()
    else:
        return "Error generating summary."

last_descriptions = []
last_contents = []

# Reset Llama memory before starting
requests.post(API_URL, headers={'Content-Type': 'application/json'}, json={
    "model": "llama3.3:70b",
    "prompt": "Forget all previous instructions and start fresh.",
    "temperature": 0.0,
    "stream": False
})

# Process each row
for i in range(1, 760):
    item_description = codif_sheet.iloc[i, description_col]
    item_content = codif_sheet.iloc[i, content_col]

    if pd.notna(item_description):
        last_descriptions.append(clean_html(item_description))
        if len(last_descriptions) > 3:
            last_descriptions.pop(0)
    
    if pd.notna(item_content):
        last_contents.append(clean_html(item_content))
        if len(last_contents) > 3:
            last_contents.pop(0)
    
    summary = generate_summary(last_descriptions, last_contents)
    print(f"📝 Row {i+1} - Generated Summary: {summary}")

    # Write summary to column F
    workbook_sheet.cell(row=i+1, column=summary_col+1, value=summary)
    workbook.save(EXCEL_FILE_PATH)

workbook.save(EXCEL_FILE_PATH)
workbook.close()

Finishing_time = datetime.now()
print("\n✅ Summaries successfully written to the Excel file.")
print(f"\n✅ Starting time: '{Starting_time}'")
print(f"\n✅ Finishing time: '{Finishing_time}'")

