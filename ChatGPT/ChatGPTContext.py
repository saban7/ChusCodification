# Codification script: Few shots (CHATGPT)
import time
import pandas as pd
import requests
import json
from openpyxl import load_workbook
from bs4 import BeautifulSoup
from difflib import get_close_matches
import time
from datetime import datetime
import os

Starting_time = datetime.now()
print(f"\n⏰ Starting time: '{Starting_time}'")

MAX_RETRIES = 3
API_URL = "https://api.openai.com/v1/chat/completions"
API_KEY = "" #API KEY

# Load the Excel file
file_path = '/home/msaban/ChusCodification/ChatGPT/FS2_GPT.xlsx'

# Load sheets using pandas
codes_sheet = pd.read_excel(file_path, sheet_name="Codes", header=None)
codif_sheet = pd.read_excel(file_path, sheet_name="Codification", header=None)

# Extract code definitions
definitions_mapping = {
    str(row[0]).strip().lower(): str(row[1]).strip() if not pd.isna(row[1]) else "No definition available"
    for row in codes_sheet.iloc[1:].values  # Skip header row
}

# Extract examples
examples_mapping = {
    str(row[0]).strip().lower(): str(row[2]).strip() if not pd.isna(row[2]) else "No example available"
    for row in codes_sheet.iloc[1:].values
}

# Get codes from "Codification" sheet 6 19 (Columns F, G → indices 5,6)
code_columns = list(range(6, 19))
codes = [str(codif_sheet.iloc[0, col]).strip().lower() for col in code_columns]

# Function to find best match
def find_best_match(code_name, available_codes):
    matches = get_close_matches(code_name, available_codes, n=1, cutoff=0.7)
    return matches[0] if matches else None

# Fix codes mapping
fixed_codes = {
    raw_code: find_best_match(raw_code, definitions_mapping.keys()) or raw_code
    for raw_code in codes
}

# 🔹 Clean HTML content
def clean_html(html_text):
    return BeautifulSoup(str(html_text), "html.parser").get_text()

# Load workbook for writing results
workbook = load_workbook(file_path)
if "Codification" not in workbook.sheetnames:
    raise ValueError("❌ Sheet 'Coding' not found in the Excel file!")

workbook_sheet = workbook["Codification"]

# Define column indices  title_col category_col   name_col   description_col      embded_col
title_col = 0       # Lesson title
category_col = 1    # Activity category
name_col = 2        # Activity name
description_col = 3 # Activity description
embed_col = 4       # Embedded media description
summary_col = 5


# Process each code column
for code_idx, code_col in enumerate(code_columns):
    raw_code_name = str(codif_sheet.iloc[0, code_col]).strip().lower()
    matched_code_name = fixed_codes.get(raw_code_name, raw_code_name)
    code_definition = definitions_mapping.get(matched_code_name, "No definition available")
    code_example = examples_mapping.get(matched_code_name, "No example available")

    print(f"\n🚀 Processing Code: '{matched_code_name}'")
    print(f"📝 Definition: {code_definition}")
    print(f"📚 Example: {code_example}\n")


    # Process each row
    for i in range(1, 90):
        
        lesson_title = codif_sheet.iloc[i, title_col]
        activity_category = codif_sheet.iloc[i, category_col]
        activity_name = codif_sheet.iloc[i, name_col]
        activity_description = codif_sheet.iloc[i, description_col]
        embed_description = codif_sheet.iloc[i, embed_col]
	previous_summary = codif_sheet.iloc[i, summary_col]
        # Build the text for the prompt conditionally    title_col category_col   name_col   description_col      embded_col
        
       # Construct the text for the prompt
        text_for_prompt = (
            f"Learning activity:\n"
            f"Activity description: {clean_html(activity_description)}.\n"
            f"Embedded media content description: {clean_html(embed_description)}.\n"
        )

        # 🔹 Construct the optimized prompt
        prompt = (
            f"You are a qualitative coding expert. You are assessing the student engagement of learning activities created by teachers in a inquiry-based learning digital platform. \n"
            f"These activities may have different media content including text and embedded artifacts (e.g., images, videos, apps, labs). Please review the provided activity description and code it based on the construct: `{matched_code_name}`. \n"
            f"The definition of this construct is `{code_definition}`.  \n"
            f"Here you have some examples: `{code_example}`. \n"
            f"For additional context, here is a summary of the 3 previous items: `{previous_summary}`. \n"    
            f"After reviewing the text, assign a code of '1' if you believe the text exemplifies `{matched_code_name}`, or a '0' if it does not.\n"
            f"Your response should only be '1' or '0', without the quotes. Do NOT provide any explanation or text after the 0 or 1. It is very important that your response is only a 0 or 1.\n\n"
            f"Text: `{text_for_prompt}`"
        )

        print(f"\n🤖 gpt-4-turbo prompt: {prompt}\n")

        # 🔹 Ensure proper JSON formatting for API request
        data = {
            "model": "gpt-4-turbo",
            "messages": [{"role": "user", "content": prompt}],  # ✅ Fixed JSON format
            "temperature": 0.0,
            "stream": False
        }

        attempt = 0
        while attempt < MAX_RETRIES:
            try:
                # Send the request to the OpenAI API
                response = requests.post(
                    API_URL,
                    headers={"Authorization": f"Bearer {API_KEY}", "Content-Type": "application/json"},
                    json=data
                )
                response.raise_for_status()  # Raise an error for HTTP issues

                # ✅ Correct API response handling
                response_json = response.json()
                api_response = response_json.get("choices", [{}])[0].get("message", {}).get("content", "").strip()

                result_value = api_response
                # Validate response format
                #if api_response and api_response[0] in ("1", "0"):
                #   result_value = api_response[0]
                #else:
                #    result_value = "Error"

                print(f"📝 Row {i+1} - Code '{matched_code_name}': API response: {result_value}")

                # Write the result to the Excel sheet
                workbook_sheet.cell(row=i+1, column=code_col+1, value=result_value)
                workbook.save(file_path)  # 🔹 Ensure changes are written to the file
                print(f"✅ Successfully written to Excel at row {i+1}, column {code_col+1}")
                print(f"\n⏳ Starting time: '{Starting_time}'")
		time.sleep(1)
                break  # Exit the retry loop if successful

            except requests.exceptions.RequestException as req_err:
                print(f"❌ API Connection Failed for row {i+1}, code '{matched_code_name}': {req_err}")
                result_value = "Error"

            except json.JSONDecodeError as json_err:
                print(f"⚠️ JSON Decode Error for row {i+1}, code '{matched_code_name}': {json_err}")
                result_value = "Error"

            # Write error result after all retries fail
            if attempt == MAX_RETRIES - 1:
                workbook_sheet.cell(row=i+1, column=code_col+1, value=result_value)
                workbook.save(file_path)  # 🔹 Save on error
                print(f"🚨 Max retries reached for row {i+1}. Moving to next item.")

            attempt += 1
            time.sleep(3)  # Wait 3s before retrying if failed

# 🔹 Ensure the workbook is properly saved and closed at the end
workbook.save(file_path)
workbook.close()
Finishing_time = datetime.now()
print("\n✅ Results successfully written to the Excel file.")
print(f"\n⏳ Starting time: '{Starting_time}'")
print(f"\n⌛ Finishing time: '{Finishing_time}'")
