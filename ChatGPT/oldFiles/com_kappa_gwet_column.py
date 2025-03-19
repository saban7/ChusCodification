import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from sklearn.metrics import cohen_kappa_score
import numpy as np

# Function to calculate GWET's AC1
def gwet_ac1(human_ratings, llama_ratings):
    n = len(human_ratings)
    if n == 0:
        return np.nan

    agreement = np.sum(human_ratings == llama_ratings) / n
    marginals_human = pd.Series(human_ratings).value_counts(normalize=True)
    marginals_llama = pd.Series(llama_ratings).value_counts(normalize=True)
    pe = np.sum(marginals_human * marginals_llama)
    ac1 = (agreement - pe) / (1 - pe) if (1 - pe) != 0 else np.nan
    return ac1

# Load the Excel file
file_path = '/Users/sabanov/Desktop/14_2_2025/Comparison/comparison.xlsx'
human = pd.read_excel(file_path, sheet_name=1, header=None)
llama = pd.read_excel(file_path, sheet_name=3, header=None)

# Load the workbook and the specific sheet
workbook = load_workbook(file_path)
comparison_sheet = workbook.worksheets[3]  # Assuming comparison sheet is the 4th sheet (index 3)

# Define the fill colors
green_fill = PatternFill(start_color="ceffce", end_color="ceffce", fill_type="solid")
red_fill = PatternFill(start_color="fd9f9f", end_color="fd9f9f", fill_type="solid")

# Define column range
start_col = 6  # Column E (0-based index for E is 4)
end_col = 18   # Column AG (0-based index for AG is 32)

# Process each column separately
for col in range(start_col, end_col + 1):  # From E to AG
    green_count = 0
    red_count = 0

    # Prepare lists for Cohen's Kappa
    human_ratings = []
    llama_ratings = []

    for row in range(1, 38):  # Rows 2 to 505 (1-based index)
        human_value = human.iloc[row, col]
        llama_value = llama.iloc[row, col]

        human_value = pd.to_numeric(human_value, errors='coerce')
        llama_value = pd.to_numeric(llama_value, errors='coerce')

        if pd.isna(human_value):
            human_value = 0

        if not pd.isna(llama_value):
            human_ratings.append(human_value)
            llama_ratings.append(llama_value)

        cell = comparison_sheet.cell(row=row + 1, column=col + 1)
        if human_value == llama_value:
            cell.fill = green_fill
            green_count += 1
        else:
            cell.fill = red_fill
            red_count += 1

    # Convert to numpy arrays
    human_ratings = np.array(human_ratings)
    llama_ratings = np.array(llama_ratings)

    # Calculate Cohen's Kappa
    kappa = cohen_kappa_score(human_ratings, llama_ratings) if len(human_ratings) > 0 else np.nan

    # Calculate GWET's AC1
    #ac1 = gwet_ac1(human_ratings, llama_ratings)

    

    # Interpret Kappa
    if pd.isna(kappa):
        interpretation = "Not enough data or insufficient unique labels"
    elif kappa < 0:
        interpretation = "Less than chance agreement"
    elif kappa <= 0.20:
        interpretation = "Slight agreement"
    elif kappa <= 0.40:
        interpretation = "Fair agreement"
    elif kappa <= 0.60:
        interpretation = "Moderate agreement"
    elif kappa <= 0.80:
        interpretation = "Substantial agreement"
    else:
        interpretation = "Almost perfect agreement"



    # Write the results to the sheet
    comparison_sheet.cell(row=40, column=col + 1).value = f"{kappa}"
    comparison_sheet.cell(row=41, column=col + 1).value = f"{interpretation}"
    #comparison_sheet.cell(row=42, column=col + 1).value = f"{ac1}"
    comparison_sheet.cell(row=43, column=col + 1).value = f"{green_count}"
    comparison_sheet.cell(row=44, column=col + 1).value = f"{red_count}"

# Save the workbook
workbook.save(file_path)
print("Processing complete. Results saved to the Excel file.")
