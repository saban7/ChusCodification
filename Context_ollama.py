import time
import pandas as pd
import requests
import json
from openpyxl import load_workbook
from bs4 import BeautifulSoup
from difflib import get_close_matches
from datetime import datetime

Starting_time = datetime.now()
print(f"\n✅ Starting time: '{Starting_time}'")

MAX_RETRIES = 3
API_URL = "http://localhost:11434/api/generate"

# Load the Excel file
file_path = '/home/msaban/ChusCodification/Codebook.xlsx'

# Load sheets using pandas
codes_sheet = pd.read_excel(file_path, sheet_name="Codes", header=None)
codif_sheet = pd.read_excel(file_path, sheet_name="Context", header=None)

# Extract code definitions
definitions_mapping = {
    str(row[0]).strip().lower(): str(row[1]).strip() if not pd.isna(row[1]) else "No definition available"
    for row in codes_sheet.iloc[1:].values  # Skip header row
}

# Extract examples
examples_mapping = {
    str(row[0]).strip().lower(): str(row[2]).strip() if not pd.isna(row[2]) else "No example available"
    for row in codes_sheet.iloc[1:].values
}

# Get codes from "Codification" sheet (Columns F, G, H → indices 5,6,7)
code_columns = list(range(5, 18))
codes = [str(codif_sheet.iloc[0, col]).strip().lower() for col in code_columns]

# Function to find best match
def find_best_match(code_name, available_codes):
    matches = get_close_matches(code_name, available_codes, n=1, cutoff=0.7)
    return matches[0] if matches else None

# Fix codes mapping
fixed_codes = {
    raw_code: find_best_match(raw_code, definitions_mapping.keys()) or raw_code
    for raw_code in codes
}

# 🔹 Clean HTML content
def clean_html(html_text):
    return BeautifulSoup(str(html_text), "html.parser").get_text()

# 🔹 Generate summary using Llama
def generate_summary(last_descriptions, last_contents):
    if not last_descriptions and not last_contents:
        return "No previous context available."
    
    summary_text = " | ".join(last_descriptions + last_contents)
    
    summary_prompt = (
        f"Summarize the following descriptions and content in a concise manner: {summary_text}"
    )
    
    data = {
        "model": "llama3.2",
        "prompt": summary_prompt,
        "temperature": 0.0,
        "stream": False
    }
    
    response = requests.post(API_URL, headers={'Content-Type': 'application/json'}, json=data)
    
    if response.status_code == 200:
        return response.json().get("response", "No summary available.").strip()
    else:
        return "Error generating summary."

# Load workbook for writing results
workbook = load_workbook(file_path)
if "Context" not in workbook.sheetnames:
    raise ValueError("❌ Sheet 'Coding' not found in the Excel file!")

workbook_sheet = workbook["Context"]

description_col = 3  # Column D
content_col = 4      # Column E

last_descriptions = []
last_contents = []

# Process each code column
for code_idx, code_col in enumerate(code_columns):
    raw_code_name = str(codif_sheet.iloc[0, code_col]).strip().lower()
    matched_code_name = fixed_codes.get(raw_code_name, raw_code_name)
    code_definition = definitions_mapping.get(matched_code_name, "No definition available")
    code_example = examples_mapping.get(matched_code_name, "No example available")

    print(f"\n🚀 Processing Code: '{matched_code_name}'")

    # Reset Ollama context
    requests.post(API_URL, headers={'Content-Type': 'application/json'}, json={
        "model": "llama3.2",
        "prompt": "Forget all previous instructions and start fresh.",
        "temperature": 0.0,
        "stream": False
    })

    for i in range(1, 284):
        item_description = codif_sheet.iloc[i, description_col]
        item_content = codif_sheet.iloc[i, content_col]

        if pd.notna(item_description):
            last_descriptions.append(clean_html(item_description))
            if len(last_descriptions) > 3:
                last_descriptions.pop(0)
        
        if pd.notna(item_content):
            last_contents.append(clean_html(item_content))
            if len(last_contents) > 3:
                last_contents.pop(0)
        
        summary = generate_summary(last_descriptions, last_contents)

        text_for_prompt = f"Item description: {clean_html(item_description)}. Item content: {clean_html(item_content)}" if pd.notna(item_description) and pd.notna(item_content) else clean_html(item_description or item_content)
        
        prompt = (
            f"Please review the provided text and code it based on the construct: `{matched_code_name}`. "
            f"The definition of this construct is `{code_definition}`. "
            f"Here you have some examples `{code_example}`. "
            f"Here you have the context: `{summary}` "
            f"After reviewing the text, assign a code of '1' if you believe the text exemplifies `{matched_code_name}`, "
            f"or a '0' if it does not. Your response should only be '1' or '0'. "
            f"Text: `{text_for_prompt}`"
        )

        data = {
            "model": "llama3.2",
            "prompt": prompt,
            "temperature": 0.0,
            "stream": False
        }

        attempt = 0
        while attempt < MAX_RETRIES:
            try:
                response = requests.post(API_URL, headers={'Content-Type': 'application/json'}, json=data)
                response.raise_for_status()
                api_response = response.json().get("response", "").strip()
                result_value = api_response[0] if api_response and api_response[0] in ("1", "0") else "Error"
                workbook_sheet.cell(row=i+1, column=code_col+1, value=result_value)
                workbook.save(file_path)
                break
            except Exception as e:
                print(f"⚠️ Error processing row {i+1}: {e}")
            attempt += 1
            time.sleep(5)

workbook.save(file_path)
workbook.close()
print("\n✅ Processing complete.")
